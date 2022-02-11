import os
import xlsxwriter
import openpyxl
import xlrd                                 #to read Excel
from openpyxl import load_workbook          #to update excel

#------------------------------------------------------------------------------------------
def GetListOfFilesFromFolder(FolderPath, FileTypeList):
        try:
                FileNameList = []
                for FileType in FileTypeList:
                        for root, dirs, files in os.walk(FolderPath):
                                for f in files:
                                        if f.endswith(FileType):
                                                FileNameList.append(os.path.join(root,f))
                                            
        except:
                print('File name collection error!\n')
        return FileNameList
#------------------------------------------------------------------------------------------
    
FileTypeList=['.py']
Path=r"D:\Spetter Network Application"
filenamelist =GetListOfFilesFromFolder (Path, FileTypeList)

#------------------Create workbook and add worksheet---------------------------------------

row = 0
col = 0
workbook = xlsxwriter.Workbook('D:\py_files.xlsx')     
worksheet = workbook.add_worksheet()
worksheet.write(0, 0, 'FileNames')
worksheet.write(0, 1, 'TPR')

#-----------------writing filesnames and TPR content into excel----------------------------
for filename in filenamelist:
        fname=filename.split("\\")
        row+=1
        worksheet.write(row, col,fname[-1])
        f1=open(filename,'r')
        line=f1.readlines()
        for ln in line:
           if ln.startswith("report.TPR"):
               newln=ln.split()
               worksheet.write(row,1,newln[2])
           else:
               pass
            
workbook.close()
f1.close()

#------------------to read excel------------------------

loc = ("D:\Exp.xlsx")
wb = xlrd.open_workbook(loc)
sheet = wb.sheet_by_index(0)
 
rownum=sheet.nrows
print rownum

excel=[]
for i in range(sheet.nrows):
    excel.append(sheet.cell_value(i,1))

#-----------------to update workbook-------------------
loc = ("D:\\py_files.xlsx")
wb = xlrd.open_workbook(loc)
sheet = wb.sheet_by_index(0)
 
rownum1=sheet.nrows

wb=load_workbook(filename='D:\\py_files.xlsx')
ws=wb.get_active_sheet()                                                 #to check sheet of opened excel
for rows in range(2,rownum+1):
    ws.cell(row=rows,column=3).value=excel[rows-1]
wb.save('D:\\py_files.xlsx')

#-------------to update py files------------------
for filename in filenamelist:
    f2=open(filename,'a+')
    line1=f2.readlines()

    wb=load_workbook(filename='D:\\py_files.xlsx')
    ws=wb.get_active_sheet()
    for j in range(1,rownum1+1):
        if filename.split("\\")[-1]==str(ws.cell(row=j,column=1).value):
            value=ws.cell(row=j,column=3).value
            break
      
    for count in range(len(line1)):
        if line1[count].startswith("report.TPR"):
                newline=line1[count].split()
                line1[count]=str(newline[0])+"\t"+"\t"+"\t"+"\t"+"\t"+str(newline[1])+" "+'"'+str(value)+'"'+"\n"
                f2=open(filename,'w')
                f2.writelines(line1)
                f2.close()
                break
print("Execution Completed")
